"""backend.app — FastAPI 应用（B0 骨架 + B1 对话端点）。

薄适配器：真正的逻辑都在 chat_service / orchestrator（可离线测）。本层只负责
HTTP/WS 收发。import-guarded —— 没装 fastapi 时本模块仍可被导入（不破坏测试）。

启动：uvicorn backend.app:app --port 9000
"""
from __future__ import annotations

from typing import Any, Dict, List, Optional

try:
    from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Depends, HTTPException, UploadFile, File, Form, Request, Response
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.middleware.cors import CORSMiddleware
    from pydantic import BaseModel
    _FASTAPI = True
except Exception:  # pragma: no cover - 沙箱无 fastapi 时走这里
    _FASTAPI = False

from secureguard.permissions import User, Role, PermissionDenied
from .state import AppState
import os as _os
# 部署入口：自动启用数据持久化（重启不丢智能体/模型/对话）。
# 落到项目根 data/ 目录，不依赖运维设环境变量。可用 DATA_DIR 覆盖位置。
_os.environ.setdefault("DATA_DIR", _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), "data"))
from .factory import build_state
from .chat_service import run_chat, collect_answer
from .auth import wechat_qr

# 全局单例 state：按环境变量选真实/内存后端（B7）。无环境变量时为全内存。
state = build_state()


if _FASTAPI:  # pragma: no cover - 需安装 fastapi 才执行
    from .deps import get_caller, get_caller_optional, get_caller_grant, require_role
    from .org_api import router as org_router
    from .kb_api import router as kb_router

    app = FastAPI(title="企业 AI 工作台 · 后端", version="0.2.0")
    app.include_router(org_router, tags=["org"])
    app.include_router(kb_router, tags=["kb"])
    app.add_middleware(
        CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"],
    )

    class ChatRequest(BaseModel):
        message: str
        agent_id: str = "general"
        session_id: str = "anon"

    class LoginRequest(BaseModel):
        phone: str
        password: str

    @app.get("/health")
    async def health() -> Dict[str, str]:
        return {"status": "ok"}

    # ---------------- 认证 ----------------
    @app.post("/v1/auth/login")
    async def login(req: "LoginRequest") -> Dict[str, Any]:
        token, msg = state.auth.login(req.phone, req.password)
        if not token:
            raise HTTPException(status_code=401, detail=msg)
        user = state.auth.resolve(token)
        return {"token": token, "user": {"id": user.id, "role": user.role.value, "dept_id": user.dept_id}}

    @app.post("/v1/auth/register")
    async def register(req: "LoginRequest") -> Dict[str, Any]:
        """自助注册。强制 role=user（防提权）；成功即返回登录态 token。"""
        token, msg = state.auth.register_self(req.phone, req.password)
        if not token:
            raise HTTPException(status_code=400, detail=msg)
        user = state.auth.resolve(token)
        return {"token": token, "user": {"id": user.id, "role": user.role.value, "dept_id": user.dept_id}}

    @app.post("/v1/auth/wechat/qr")
    async def wechat_login_qr() -> Dict[str, str]:
        return wechat_qr()  # 桩：未配置时返回 not_configured

    @app.get("/v1/me")
    async def me(user: "User" = Depends(get_caller)) -> Dict[str, Any]:
        return {"id": user.id, "role": user.role.value, "dept_id": user.dept_id}

    # ---------------- 用户批量导入（管理员） ----------------
    from backend.user_bulk_import import import_users, template_csv, IMPORTABLE_ROLES

    @app.get("/v1/users/import/template")
    async def user_import_template_ep(_: "User" = Depends(require_role(Role.ADMIN))):
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse(template_csv(), media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=org_import_template.csv"})

    @app.post("/v1/users/import")
    async def user_import_ep(file: "UploadFile" = File(...),
                             caller: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        data = await file.read()
        name = (file.filename or "").lower()
        rows = []
        if name.endswith((".xlsx", ".xls")):
            try:
                import openpyxl, io as _io
                wb = openpyxl.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
                ws = wb.active
                it = ws.iter_rows(values_only=True)
                headers = [str(h).strip() if h is not None else "" for h in next(it, [])]
                for r in it:
                    rows.append({headers[i]: ("" if v is None else str(v)) for i, v in enumerate(r) if i < len(headers)})
            except ImportError:
                raise HTTPException(status_code=400, detail="服务器缺 openpyxl,请用 CSV 或安装 openpyxl")
        else:
            import csv as _csv, io as _io
            text = data.decode("utf-8", errors="replace")
            rows = [r for r in _csv.DictReader(_io.StringIO(text)) if not (r.get("phone","") or "").startswith("#")]

        existing_phones = {u["phone"] for u in state.auth.list_users()}
        def _exists(p): return p in existing_phones
        def _register(phone, pwd, role, dept):
            uid = state.auth.register(phone, pwd, role, dept)
            existing_phones.add(phone)
            return uid

        rep = import_users(rows, register_fn=_register, exists_fn=_exists)
        return rep.public()

    # ---------------- RBAC 保护示例（admin 专属） ----------------
    @app.get("/v1/audit/verify")
    async def audit_verify(_: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        return state.auditor.verify_chain()

    # ---------------- 组织权限管理（org_core 显形——v0.8.0）----------------
    from backend.org_admin import build_tree, list_users as org_list_users, list_roles, flat_nodes
    from org_core import NodeType, PermissionDenied
    from org_core.permissions import validate_perm_keys

    def _org(ident):
        """从身份取 (repo, svc, tenant, grant)。"""
        return state.identity.repo, state.identity.svc, ident.tenant_id, ident.grant_id

    @app.get("/v1/org/tree")
    async def org_tree_ep(ident = Depends(get_caller_grant)) -> Dict[str, Any]:
        repo = state.identity.repo
        return build_tree(repo, ident.tenant_id, state.identity.root_id) or {}

    @app.get("/v1/org/users")
    async def org_users_ep(ident = Depends(get_caller_grant)) -> list:
        try:
            return org_list_users(state.identity.repo, ident.tenant_id)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"org/users: {type(e).__name__}: {e}")

    @app.get("/v1/org/roles")
    async def org_roles_ep(ident = Depends(get_caller_grant)) -> list:
        return list_roles(state.identity.repo, ident.tenant_id)

    @app.get("/v1/org/nodes")
    async def org_nodes_ep(ident = Depends(get_caller_grant)) -> list:
        return flat_nodes(state.identity.repo, ident.tenant_id)

    class NodeCreate(BaseModel):
        parent_id: str; name: str; type: str = "department"

    @app.post("/v1/org/node")
    async def org_create_node_ep(req: "NodeCreate", ident = Depends(get_caller_grant)) -> Dict[str, Any]:
        repo, svc, tenant, grant = _org(ident)
        try:
            n = svc.create_node(tenant, grant, parent_id=req.parent_id, type=NodeType(req.type), name=req.name)
            return {"id": n.id, "name": n.name, "type": n.type.value, "parent_id": n.parent_id}
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))

    class NodeRename(BaseModel):
        name: str

    @app.post("/v1/org/node/{node_id}/rename")
    async def org_rename_node_ep(node_id: str, req: "NodeRename", ident = Depends(get_caller_grant)) -> Dict[str, Any]:
        repo, svc, tenant, grant = _org(ident)
        try:
            n = svc.rename_node(tenant, grant, node_id, req.name)
            return {"id": n.id, "name": n.name}
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))

    class RoleCreate(BaseModel):
        at_node_id: str; name: str; perm_keys: list

    @app.post("/v1/org/role")
    async def org_define_role_ep(req: "RoleCreate", ident = Depends(get_caller_grant)) -> Dict[str, Any]:
        repo, svc, tenant, grant = _org(ident)
        try:
            r = svc.define_role(tenant, grant, req.at_node_id, req.name, validate_perm_keys(frozenset(req.perm_keys)))
            return {"id": r.id, "name": r.name, "perms": sorted(r.perm_keys)}
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    class GrantReq(BaseModel):
        user_id: str; role_id: str; org_node_id: str; label: str = ""

    @app.post("/v1/org/grant")
    async def org_grant_ep(req: "GrantReq", ident = Depends(get_caller_grant)) -> Dict[str, Any]:
        repo, svc, tenant, grant = _org(ident)
        try:
            # #7 审计守卫：审计员只能由平台管理员任命
            from backend.audit_appointment_guard import guard_appoint
            role_name = next((r["name"] for r in list_roles(repo, tenant)
                              if r["id"] == req.role_id), "")
            guard_appoint(repo, tenant, ident.person_id, state.identity.root_id, role_name)
            g = svc.grant_role(tenant, grant, user_id=req.user_id, role_id=req.role_id,
                               org_node_id=req.org_node_id, label=req.label)
            return {"id": g.id, "label": g.label}
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))

    class AdminScopeReq(BaseModel):
        user_id: str; org_node_id: str

    @app.post("/v1/org/admin-scope")
    async def org_admin_scope_ep(req: "AdminScopeReq", ident = Depends(get_caller_grant)) -> Dict[str, Any]:
        repo, svc, tenant, grant = _org(ident)
        try:
            s = svc.grant_admin_scope(tenant, grant, user_id=req.user_id, org_node_id=req.org_node_id)
            return {"id": s.id}
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))

    @app.get("/v1/org/perms")
    async def org_perms_ep(_ = Depends(get_caller_grant)) -> list:
        from org_core.permissions import ALL_PERMS
        return sorted(ALL_PERMS)

    # ---------------- 多级组织架构 Excel 导入（v0.8.1）----------------
    from backend.org_excel_import import import_org

    def _read_sheet(wb, names):
        for nm in names:
            if nm in wb.sheetnames:
                ws = wb[nm]; it = ws.iter_rows(values_only=True)
                headers = [str(h).strip() if h is not None else "" for h in next(it, [])]
                return [{headers[i]: ("" if v is None else str(v)) for i, v in enumerate(r) if i < len(headers)} for r in it]
        return []

    @app.post("/v1/org/import")
    async def org_import_ep(file: "UploadFile" = File(...),
                            dry_run: bool = Form(False),
                            ident = Depends(get_caller_grant)) -> Dict[str, Any]:
        try:
            import openpyxl, io as _io
        except ImportError:
            raise HTTPException(status_code=400, detail="服务器缺 openpyxl")
        data = await file.read()
        try:
            wb = openpyxl.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
        except Exception:
            raise HTTPException(status_code=400, detail="无法读取 Excel")
        org_rows = _read_sheet(wb, ["组织", "组织表", "org", "Sheet1"])
        people_rows = _read_sheet(wb, ["人员", "人员表", "people", "Sheet2"])
        existing = {u["phone"] for u in state.auth.list_users()}
        rep = import_org(
            org_rows, people_rows,
            svc=state.identity.svc, repo=state.identity.repo,
            tenant_id=ident.tenant_id, root_id=state.identity.root_id,
            importer_grant_id=ident.grant_id,
            register_fn=state.auth.register,
            provision_person=state.identity.provision_person,
            exists_phone=lambda p: p in existing,
            dry_run=dry_run,
        )
        return rep.public()

    @app.get("/v1/org/import/template")
    async def org_import_template_ep(_ = Depends(get_caller_grant)):
        try:
            import openpyxl, io as _io
            from fastapi.responses import StreamingResponse
            wb = openpyxl.Workbook()
            ws1 = wb.active; ws1.title = "组织"
            ws1.append(["节点名", "上级节点", "类型"])
            ws1.append(["华东子公司", "总部", "子公司"])
            ws1.append(["机电部", "华东子公司", "部门"])
            ws1.append(["A项目部", "机电部", "项目部"])
            ws2 = wb.create_sheet("人员")
            ws2.append(["手机号", "姓名", "岗位", "所属节点", "初始密码"])
            ws2.append(["13800001111", "张三", "项目经理", "A项目部", ""])
            ws2.append(["13800001111", "张三", "成员", "机电部", ""])
            buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
            return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=org_template.xlsx"})
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"模板生成失败: {type(e).__name__}: {e}")

    # ---------------- Coze SSO（v0.8.4）----------------
    import os as _os
    from backend.coze_sso import CozeSSO

    @app.post("/v1/coze/sso")
    async def coze_sso_ep(ident = Depends(get_caller_grant), resp: Response = None):
        COZE_BASE = _os.getenv("COZE_BASE_URL", "http://127.0.0.1:8888")
        COZE_SECRET = _os.getenv("COZE_SSO_SECRET", "change-me-please")
        sso = CozeSSO(COZE_BASE, COZE_SECRET)
        session_key = sso.mint_session(ident.user_id)
        if session_key is None:
            raise HTTPException(status_code=502, detail="Coze SSO 失败")
        if resp:
            resp.set_cookie("session_key", session_key, path="/", httponly=True, samesite="lax")
        return {"ok": True}

    # ---------------- 智能体 CRUD + 发布状态机（B3） ----------------
    def _view(rec: Dict[str, Any]) -> Dict[str, Any]:
        """对外视图（含 status/visibility，便于前端区分草稿/已发布）。"""
        return {k: rec[k] for k in (
            "id", "name", "description", "domain", "icon", "visibility", "status",
            "tools_count", "skills_count", "kb_count", "free_quota_tokens",
        ) if k in rec}

    @app.get("/v1/agents")
    async def list_agents(caller: "Optional[User]" = Depends(get_caller_optional)) -> list:
        return [_view(r) for r in state.agent_service.list_for(caller)]

    @app.get("/v1/agents/{agent_id}")
    async def get_agent(agent_id: str,
                        caller: "Optional[User]" = Depends(get_caller_optional)) -> Dict[str, Any]:
        rec = state.agent_service.get(agent_id)
        if not rec:
            raise HTTPException(status_code=404, detail="智能体不存在")
        from secureguard.permissions import can_use_agent
        from .agent_service import _ANON
        if not can_use_agent(caller or _ANON, state.agent_service._rbac(rec))[0]:
            raise HTTPException(status_code=403, detail="无权访问该智能体")
        return _view(rec)

    class AgentCreate(BaseModel):
        name: str
        visibility: str = "private"
        domain: str = "general"
        scope: str = "open"
        description: str = ""
        dept_id: Optional[str] = None

    # ---------------- 智能体 MD 导入/导出 ----------------
    class AgentMdImport(BaseModel):
        content: str

    @app.post("/v1/agents/import-md")
    async def import_agent_md_ep(req: "AgentMdImport",
                                 caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        from backend.agent_md import parse_agent_md
        try:
            payload = parse_agent_md(req.content)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"MD 解析失败：{e}")
        try:
            rec = state.agent_service.create(caller, payload)
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))
        return _view(rec)

    @app.get("/v1/agents/{agent_id}/export-md")
    async def export_agent_md_ep(agent_id: str,
                                 caller: "Optional[User]" = Depends(get_caller_optional)):
        from backend.agent_md import to_agent_md
        rec = state.agent_service.get(agent_id) if hasattr(state.agent_service, "get") else None
        if rec is None:
            for a in state.agent_service.list_for(caller):
                if a.get("id") == agent_id:
                    rec = a
                    break
        if rec is None:
            raise HTTPException(status_code=404, detail="智能体不存在")
        md = to_agent_md(rec, rec.get("system_prompt", ""))
        return {"id": agent_id, "filename": f"{rec.get('name', agent_id)}.md", "content": md}

    # ---- zip 批量导入智能体（v0.6.0） ----
    @app.post("/v1/agents/import-zip")
    async def import_agents_zip_ep(file: "UploadFile" = File(...),
                                   conflict: str = Form("skip"),
                                   caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        from backend.agent_md import parse_agent_md
        from backend.agent_batch_import import import_zip
        data = await file.read()
        existing = {a.get("name") for a in state.agent_service.list_for(caller)}

        def _exists(n): return n in existing
        def _create(payload):
            rec = state.agent_service.create(caller, payload)
            existing.add(rec.get("name"))
            return rec
        def _delete(n):
            if hasattr(state.agent_service, "delete_by_name"):
                state.agent_service.delete_by_name(caller, n)
            existing.discard(n)

        rep = import_zip(data, parse_fn=parse_agent_md, create_fn=_create,
                         exists_fn=_exists, delete_fn=_delete, conflict=conflict)
        return rep.public()

    @app.post("/v1/admin/agents")
    async def create_agent_ep(req: "AgentCreate",
                              caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        try:
            rec = state.agent_service.create(caller, req.model_dump())
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))
        return _view(rec)

    def _lifecycle(action: str):
        async def _ep(agent_id: str, caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
            fn = {"submit": state.agent_service.submit, "approve": state.agent_service.approve,
                  "reject": state.agent_service.reject}[action]
            try:
                return _view(fn(caller, agent_id))
            except KeyError:
                raise HTTPException(status_code=404, detail="智能体不存在")
            except PermissionDenied as e:
                raise HTTPException(status_code=403, detail=str(e))
        return _ep

    app.post("/v1/admin/agents/{agent_id}/submit")(_lifecycle("submit"))
    app.post("/v1/admin/agents/{agent_id}/approve")(_lifecycle("approve"))
    app.post("/v1/admin/agents/{agent_id}/reject")(_lifecycle("reject"))

    # ---------------- 配额（B4） ----------------
    @app.get("/v1/quota/{agent_id}")
    async def quota_status(agent_id: str, caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        agent = state.agent_service.get(agent_id) or {}
        free = int(agent.get("free_quota_tokens", 10000))
        return await state.quota.status((caller.id, agent_id), limit=free)

    @app.post("/v1/quota/{agent_id}/recharge")
    async def quota_recharge(agent_id: str, add_tokens: int,
                             caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        """用户为自己的某智能体充值（立即解锁）。"""
        return await state.quota.recharge((caller.id, agent_id), add_tokens)

    class QuotaAdjust(BaseModel):
        user_id: str
        agent_id: str
        limit: int

    @app.post("/v1/admin/quota")
    async def quota_adjust(req: "QuotaAdjust", caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        """管理员调整某用户配额。权限走 permissions.can_manage_quota（admin 全局/部门管理员限本部门）。"""
        from secureguard.permissions import can_manage_quota
        target = state.auth.get_user(req.user_id)
        if target is None:
            raise HTTPException(status_code=404, detail="目标用户不存在")
        ok, reason = can_manage_quota(caller, target)
        if not ok:
            raise HTTPException(status_code=403, detail=reason)
        return await state.quota.set_limit((req.user_id, req.agent_id), req.limit)

    # ---------------- 审计列表（admin） ----------------
    @app.get("/v1/audit")
    async def audit_list(_: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        recent = [
            {"hash": e.entry_hash[:7], "stage": e.stage, "decision": e.decision,
             "time": e.timestamp[11:19]}
            for e in state.auditor.entries[-10:]
        ]
        return {"summary": state.auditor.summary(), "recent": recent}

    # ---------------- 知识库（B5：隔离检索） ----------------
    @app.get("/v1/knowledge")
    async def list_knowledge(caller: "Optional[User]" = Depends(get_caller_optional)) -> list:
        return state.kb_service.list_for(caller)

    class KBSearch(BaseModel):
        query: str
        top_k: int = 5

    @app.post("/v1/kb/search")
    async def kb_search(req: "KBSearch",
                        caller: "Optional[User]" = Depends(get_caller_optional)) -> Dict[str, Any]:
        # 隔离下推：只在调用者可达的库里检索
        return state.kb_service.search(caller, req.query, req.top_k)

    # ---------------- 会话历史（持久化 + 搜索） ----------------
    @app.get("/v1/sessions")
    async def list_sessions(caller: "User" = Depends(get_caller)) -> list:
        return state.sessions.list_sessions(user_id=caller.id)

    @app.get("/v1/sessions/{sid}/messages")
    async def session_messages(sid: str, caller: "User" = Depends(get_caller)) -> list:
        return [m.to_dict() for m in state.sessions.messages(sid)]

    @app.get("/v1/sessions/search")
    async def search_sessions(q: str, caller: "User" = Depends(get_caller)) -> list:
        return state.sessions.search(caller.id, q)

    class RenameReq(BaseModel):
        title: str

    @app.post("/v1/sessions/{sid}/rename")
    async def rename_session(sid: str, req: "RenameReq", caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        ok = state.sessions.rename(sid, req.title)
        if not ok:
            raise HTTPException(status_code=404, detail="会话不存在")
        return {"id": sid, "title": req.title}

    @app.post("/v1/sessions/{sid}/delete")
    async def delete_session(sid: str, caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        return {"deleted": state.sessions.delete(sid)}

    # ---------------- 评测任务集 + 运行器（进化燃料，admin）----------------
    @app.get("/v1/admin/eval/tasks")
    async def eval_tasks_ep(_: "User" = Depends(require_role(Role.ADMIN))) -> list:
        from backend.eval_runner import default_tasks
        from dataclasses import asdict
        return [asdict(t) for t in default_tasks()]

    @app.post("/v1/admin/eval/run")
    async def eval_run_ep(caller: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        """用当前上线的思考 MD 跑默认评测集，返回失败聚类 + 是否触发进化。"""
        from backend.eval_runner import EvalRunner, default_tasks
        from backend.sandbox_executor import build_executor
        from backend.chat_service import run_chat, collect_answer
        import asyncio
        tasks = default_tasks()
        executor = build_executor()
        runner = EvalRunner(executor=executor)

        def answer_fn(task):
            # 用平台真实链路产生回答（走当前上线的思考 MD）
            agent_id = "general" if task.domain != "software" else "general"
            res = asyncio.get_event_loop().run_until_complete(
                collect_answer(run_chat(state, task.prompt, agent_id=agent_id,
                                        session_id=f"eval_{task.id}",
                                        caller={"id": caller.id, "role": "admin"})))
            return res.get("answer", "")

        run = runner.run(tasks, answer_fn, md_name="main")
        triggers = EvalRunner.should_trigger_evolution(run, threshold=3)
        out = run.public()
        out["evolution_triggers"] = triggers
        return out

    # ---------------- 思考 MD 版本管理（进化系统地基，admin）----------------
    # 端点走 /v1/admin*，Caddy 已有，无需新增规则。
    @app.get("/v1/admin/md/names")
    async def md_names_ep(_: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        return {"names": state.md_versions.list_md_names()}

    @app.get("/v1/admin/md/{md_name}/versions")
    async def md_versions_ep(md_name: str, _: "User" = Depends(require_role(Role.ADMIN))) -> list:
        return state.md_versions.list_versions(md_name)

    @app.get("/v1/admin/md/version/{version_id}")
    async def md_get_version_ep(version_id: str, _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        v = state.md_versions.get_version(version_id)
        if not v:
            raise HTTPException(status_code=404, detail="版本不存在")
        v["content"] = state.md_versions.get_content(version_id)
        return v

    class MdSaveReq(BaseModel):
        md_name: str
        content: str
        note: str = ""

    @app.post("/v1/admin/md/save")
    async def md_save_ep(req: "MdSaveReq", caller: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        return state.md_versions.save_version(req.md_name, req.content, origin="manual",
                                              created_by=caller.id, note=req.note)

    @app.post("/v1/admin/md/version/{version_id}/set-live")
    async def md_set_live_ep(version_id: str, caller: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        try:
            return state.md_versions.set_live(version_id, by=caller.id)
        except KeyError as e:
            raise HTTPException(status_code=404, detail=str(e))

    @app.post("/v1/admin/md/{md_name}/rollback/{version_id}")
    async def md_rollback_ep(md_name: str, version_id: str, caller: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        try:
            return state.md_versions.rollback(md_name, version_id, by=caller.id)
        except KeyError as e:
            raise HTTPException(status_code=404, detail=str(e))

    @app.get("/v1/admin/md/diff")
    async def md_diff_ep(a: str, b: str, _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        try:
            return state.md_versions.diff(a, b)
        except KeyError as e:
            raise HTTPException(status_code=404, detail=str(e))

    @app.get("/v1/admin/permissions")
    async def permission_handbook_ep(_: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        from backend.permission_matrix import build_handbook
        return build_handbook()

    @app.get("/v1/me/capabilities")
    async def my_capabilities_ep(caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        from backend.permission_matrix import capabilities_of
        return {"role": caller.role.value, "capabilities": capabilities_of(caller.role)}

    # ---------------- sk 出站调用密钥（管理） ----------------
    # 📌 Caddy 需加： /v1/sk*  → localhost:9000
    class SkIssue(BaseModel):
        label: str = ""
        agent_ids: List[str] = []
        meter: str = "calls"               # calls | tokens
        limit: Optional[int] = None        # 额度上限（None=不限）
        reset_mode: str = "none"           # none | period
        period: str = "monthly"
        allow_recharge: bool = True

    @app.post("/v1/sk/issue")
    async def sk_issue_ep(req: "SkIssue", caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        from backend.permission_matrix import can, CAP_ISSUE_SK
        if not can(caller.role, CAP_ISSUE_SK):
            raise HTTPException(status_code=403, detail="无权签发调用密钥（仅超级管理员）")
        try:
            return state.sk.issue(owner_id=caller.id, label=req.label, agent_ids=req.agent_ids,
                                  meter=req.meter, limit=req.limit, reset_mode=req.reset_mode,
                                  period=req.period, allow_recharge=req.allow_recharge)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    @app.get("/v1/sk/list")
    async def sk_list_ep(caller: "User" = Depends(get_caller)) -> list:
        return state.sk.list_for_owner(caller.id)

    @app.post("/v1/sk/{kid}/revoke")
    async def sk_revoke_ep(kid: str, caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        try:
            return {"revoked": state.sk.revoke(kid, owner_id=caller.id)}
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))

    class SkRecharge(BaseModel):
        add: int

    @app.post("/v1/sk/{kid}/recharge")
    async def sk_recharge_ep(kid: str, req: "SkRecharge", caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        try:
            return state.sk.recharge(kid, req.add, owner_id=caller.id)
        except (KeyError, ValueError) as e:
            raise HTTPException(status_code=400, detail=str(e))
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))

    # ---------------- OpenAI 兼容被调接口（外部平台用 sk- 调你的智能体） ----------------
    # 📌 Caddy 需加： /v1/openai*  → localhost:9000
    # 验 sk → 查 scope → 查配额 → 走完整安全内核(L0-L4) → 应答 → 扣量 → 超额 429
    @app.post("/v1/openai/chat/completions")
    async def openai_compat_ep(req: "Request") -> Any:  # type: ignore
        from backend.chat_service import run_chat, collect_answer
        # 1) 取 sk（Authorization: Bearer sk-xxx）
        auth = req.headers.get("authorization", "")
        raw_key = auth[7:].strip() if auth.lower().startswith("bearer ") else ""
        rec = state.sk.verify(raw_key)
        if rec is None:
            raise HTTPException(status_code=401, detail="无效或已吊销的 sk")
        body = await req.json()
        # OpenAI 请求：model 字段我们用作 agent_id；messages 取最后一条 user
        agent_id = body.get("model", "general")
        messages = body.get("messages", [])
        user_msg = ""
        for m in reversed(messages):
            if m.get("role") == "user":
                user_msg = m.get("content", "")
                break
        # 2) scope 检查
        if not state.sk.can_call_agent(rec, agent_id):
            raise HTTPException(status_code=403, detail=f"此 sk 无权调用智能体 {agent_id}")
        # 3) 配额检查
        q = state.sk.check_quota(rec)
        if not q["ok"]:
            raise HTTPException(status_code=429, detail=f"额度已用尽（{rec.meter}）。请续费或更换密钥。")
        # 4) 走完整安全内核 L0-L4（与站内对话同一条 run_chat，安全门控照常生效）
        sk_caller = {"id": f"sk:{rec.id}", "role": "user", "via_sk": rec.id}
        result = await collect_answer(
            run_chat(state, user_msg, agent_id=agent_id, session_id=f"sk_{rec.id}",
                     caller=sk_caller))
        answer = result.get("answer", "")
        blocked = result.get("blocked", False)
        # 5) 计量扣减（calls 或 tokens）
        est_tokens = max(1, (len(user_msg) + len(answer)) // 2)
        state.sk.consume(rec, calls=1, tokens=est_tokens)
        # 6) 审计
        state.auditor.log_stage(stage="SK_CALL", decision="BLOCKED" if blocked else "PASS",
                                session_id=f"sk_{rec.id}",
                                extra={"sk_id": rec.id, "agent_id": agent_id,
                                       "meter": rec.meter, "tokens": est_tokens})
        # 7) 返回 OpenAI 兼容格式
        return {
            "id": f"chatcmpl-{rec.id}", "object": "chat.completion",
            "created": int(__import__("time").time()), "model": agent_id,
            "choices": [{"index": 0, "finish_reason": "stop",
                         "message": {"role": "assistant", "content": answer}}],
            "usage": {"prompt_tokens": len(user_msg) // 2,
                      "completion_tokens": len(answer) // 2, "total_tokens": est_tokens},
        }

    # ---------------- 文件上传（②） ----------------
    # 📌 Caddy 需加： /v1/files*  → localhost:9000
    @app.post("/v1/files/upload")
    async def upload_file_ep(file: "UploadFile" = File(...),  # type: ignore
                             session_id: str = Form("anon"),
                             caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        data = await file.read()
        try:
            return state.files.ingest(file.filename, data, session_id=session_id)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    @app.get("/v1/files/session/{sid}")
    async def list_session_files_ep(sid: str, caller: "User" = Depends(get_caller)) -> list:
        return state.files.list_for_session(sid)

    @app.post("/v1/files/{file_id}/delete")
    async def delete_file_ep(file_id: str, caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        return {"deleted": state.files.delete(file_id)}

    # ---------------- 语音 ASR/TTS（③④⑤） ----------------
    # 📌 Caddy 需加： /v1/voice*  → localhost:9000
    @app.get("/v1/voice/config")
    async def voice_config_ep(caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        return state.voice.config()

    @app.post("/v1/voice/asr")
    async def voice_asr_ep(file: "UploadFile" = File(...),  # type: ignore
                           caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        audio = await file.read()
        return state.voice.asr.transcribe(audio)

    class TTSReq(BaseModel):
        text: str
        voice: Optional[str] = None
        rate: float = 1.0
        pitch: float = 1.0

    @app.post("/v1/voice/tts")
    async def voice_tts_ep(req: "TTSReq", caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        return state.voice.tts.synthesize(req.text, req.voice, req.rate, req.pitch)

    # ---------------- 模型配置（① 选模型 + 管理员配置） ----------------
    # 📌 Caddy 需加： /v1/models*  → localhost:9000 （放在 /v1/* 通配之前）
    @app.get("/v1/models")
    async def list_models_ep(caller: "User" = Depends(get_caller)) -> list:
        # 聊天框下拉：所有登录用户可见已启用模型
        return state.models.list_selectable()

    @app.get("/v1/models/all")
    async def list_all_models_ep(_: "User" = Depends(require_role(Role.ADMIN))) -> list:
        return state.models.list_models(include_disabled=True)

    @app.get("/v1/models/providers")
    async def model_providers_ep(_: "User" = Depends(require_role(Role.ADMIN))) -> list:
        from backend.model_adapter import list_providers
        return list_providers()

    class FetchModelsReq(BaseModel):
        api_base: str
        api_key: str

    @app.post("/v1/models/fetch")
    async def fetch_models_ep(req: "FetchModelsReq",
                              _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        from backend.model_adapter import fetch_models
        return fetch_models(req.api_base, req.api_key)

    class ModelCreate(BaseModel):
        name: str
        api_base: str = ""
        api_key: str = ""
        model: str
        enabled: bool = True

    @app.post("/v1/models")
    async def create_model_ep(req: "ModelCreate",
                              _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        try:
            return state.models.create(req.name, req.api_base, req.api_key, req.model, req.enabled)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    class ModelUpdate(BaseModel):
        name: Optional[str] = None
        api_base: Optional[str] = None
        api_key: Optional[str] = None
        model: Optional[str] = None
        enabled: Optional[bool] = None

    @app.post("/v1/models/{model_id}/update")
    async def update_model_ep(model_id: str, req: "ModelUpdate",
                              _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        try:
            return state.models.update(model_id, **req.model_dump())
        except KeyError:
            raise HTTPException(status_code=404, detail="模型不存在")
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    @app.post("/v1/models/{model_id}/delete")
    async def delete_model_ep(model_id: str,
                              _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        try:
            return {"deleted": state.models.delete(model_id)}
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    # ---------------- 算力 / 管理统计（B6） ----------------
    from .admin_service import compute_status, admin_stats

    @app.get("/v1/compute/status")
    async def compute_status_ep() -> list:
        return compute_status()

    @app.get("/v1/admin/stats")
    async def admin_stats_ep(_: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        return admin_stats(state)

    # ---------------- 用户管理（admin） ----------------
    @app.get("/v1/admin/users")
    async def list_users(caller: "User" = Depends(get_caller)) -> list:
        try:
            return state.user_admin.list_users(caller)
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))

    class AdminUserCreate(BaseModel):
        phone: str
        password: Optional[str] = None
        role: str = "user"
        dept_id: Optional[str] = None

    @app.post("/v1/admin/users")
    async def admin_create_user(req: "AdminUserCreate", caller: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        import re, secrets
        phone = (req.phone or "").strip()
        if not re.fullmatch(r"1[3-9]\d{9}", phone):
            raise HTTPException(status_code=400, detail="手机号格式不正确")
        if any(u["phone"] == phone for u in state.auth.list_users()):
            raise HTTPException(status_code=400, detail="该手机号已存在")
        valid = {"user", "developer", "department_admin", "auditor", "admin"}
        if req.role not in valid:
            raise HTTPException(status_code=400, detail=f"无效角色:{req.role}")
        pwd = (req.password or "").strip() or ("Aa" + secrets.token_hex(3))
        uid = state.auth.register(phone, pwd, req.role, req.dept_id)
        return {"id": uid, "phone": phone, "role": req.role, "dept_id": req.dept_id, "init_password": pwd}

    class RoleUpdate(BaseModel):
        role: str

    @app.post("/v1/admin/users/{user_id}/role")
    async def set_user_role(user_id: str, req: "RoleUpdate",
                            caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        try:
            return state.user_admin.set_role(caller, user_id, req.role)
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))
        except KeyError:
            raise HTTPException(status_code=404, detail="用户不存在")
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    class DeptUpdate(BaseModel):
        dept_id: Optional[str] = None

    @app.post("/v1/admin/users/{user_id}/department")
    async def assign_department(user_id: str, req: "DeptUpdate",
                                caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        try:
            return state.user_admin.assign_department(caller, user_id, req.dept_id)
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))
        except KeyError:
            raise HTTPException(status_code=404, detail="用户不存在")

    @app.post("/v1/admin/users/{user_id}/delete")
    async def delete_user(user_id: str, caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        try:
            return state.user_admin.delete_user(caller, user_id)
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))
        except KeyError:
            raise HTTPException(status_code=404, detail="用户不存在")

    # ---------------- 配额策略（可组合，admin/受限管理员设） ----------------
    class QuotaPolicyReq(BaseModel):
        agent_id: str
        reset_mode: str = "period"        # period | cooldown | none
        period: str = "monthly"           # monthly | weekly
        cooldown_seconds: int = 3600
        allow_recharge: bool = True

    @app.post("/v1/admin/quota/policy")
    async def set_quota_policy(req: "QuotaPolicyReq", caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        from secureguard.permissions import can_manage_quota
        # 复用配额管理权（admin/受限管理员）；target 用 caller 占位（策略是按智能体的全局设置）
        ok, why = can_manage_quota(caller, caller)
        if not ok:
            raise HTTPException(status_code=403, detail=why)
        if req.reset_mode not in ("period", "cooldown", "none"):
            raise HTTPException(status_code=400, detail="reset_mode 必须是 period/cooldown/none")
        from backend.quota_service import QuotaPolicy
        pol = QuotaPolicy(reset_mode=req.reset_mode, period=req.period,
                          cooldown_seconds=req.cooldown_seconds, allow_recharge=req.allow_recharge)
        state.quota.set_policy(req.agent_id, pol)
        return {"agent_id": req.agent_id, "policy": pol.to_dict()}

    @app.get("/v1/admin/quota/policy/{agent_id}")
    async def get_quota_policy(agent_id: str, caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        from secureguard.permissions import can_manage_quota
        ok, why = can_manage_quota(caller, caller)
        if not ok:
            raise HTTPException(status_code=403, detail=why)
        pol = state.quota.get_policy(agent_id)
        return {"agent_id": agent_id, "policy": pol.to_dict() if pol else None}

    # ---------------- 部门申请（用户提交 / admin 审批） ----------------
    class DeptRequest(BaseModel):
        dept_id: str

    @app.post("/v1/dept/request")
    async def request_department(req: "DeptRequest", caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        return state.user_admin.request_department(caller, req.dept_id)

    @app.get("/v1/admin/dept/requests")
    async def list_dept_requests(caller: "User" = Depends(get_caller)) -> list:
        try:
            return state.user_admin.list_requests(caller)
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))

    @app.post("/v1/admin/dept/requests/{req_id}/{action}")
    async def handle_dept_request(req_id: str, action: str,
                                  caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        if action not in ("approve", "reject"):
            raise HTTPException(status_code=400, detail="action 必须是 approve/reject")
        try:
            fn = state.user_admin.approve_request if action == "approve" else state.user_admin.reject_request
            return fn(caller, req_id)
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))
        except KeyError:
            raise HTTPException(status_code=404, detail="申请不存在")

    # ---------------- 受控审计读取（仅 auditor，必留痕） ----------------
    @app.get("/v1/audit/kb/list")
    async def audit_kb_list(caller: "User" = Depends(get_caller)) -> list:
        try:
            return state.kb_service.audit_list_all(caller)
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))

    class AuditRead(BaseModel):
        kb_id: str
        reason: str

    @app.post("/v1/audit/kb/read")
    async def audit_kb_read(req: "AuditRead", caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        try:
            return state.kb_service.audit_read(caller, req.kb_id, req.reason, state.auditor)
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))
        except KeyError:
            raise HTTPException(status_code=404, detail="知识库不存在")
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    # ---------------- 知识库入库（建库/粘贴文本/上传文档） ----------------
    import secrets as _secrets

    class KBCreate(BaseModel):
        name: str
        visibility: str = "private"
        dept_id: Optional[str] = None

    @app.post("/v1/kb/create")
    async def create_kb_ep(req: "KBCreate", caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        try:
            return state.kb_service.create_kb(caller, req.name, req.visibility, req.dept_id)
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    class KBIngestText(BaseModel):
        text: str
        doc_id: Optional[str] = None

    @app.post("/v1/kb/{kb_id}/ingest")
    async def ingest_kb_ep(kb_id: str, req: "KBIngestText",
                           caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        if not req.text.strip():
            raise HTTPException(status_code=400, detail="文本为空")
        try:
            return state.kb_service.ingest_text(
                caller, kb_id, req.doc_id or ("doc_" + _secrets.token_hex(4)), req.text)
        except KeyError:
            raise HTTPException(status_code=404, detail="知识库不存在")
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))

    @app.post("/v1/kb/{kb_id}/upload")
    async def upload_kb_doc_ep(kb_id: str, file: "UploadFile" = File(...),
                               caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        data = await file.read()
        try:
            text = data.decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="当前仅支持 txt/md 文本;PDF/Word 解析待接")
        try:
            return state.kb_service.ingest_text(caller, kb_id, file.filename or "upload", text)
        except KeyError:
            raise HTTPException(status_code=404, detail="知识库不存在")
        except PermissionDenied as e:
            raise HTTPException(status_code=403, detail=str(e))

    class KbVisibility(BaseModel):
        visibility: str

    @app.post("/v1/kb/{kb_id}/visibility")
    async def set_kb_visibility_ep(kb_id: str, req: "KbVisibility",
                                   caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        if req.visibility not in {"private", "department", "public"}:
            raise HTTPException(status_code=400, detail="非法可见性")
        return {"ok": bool(state.kb_service.set_visibility(caller, kb_id, req.visibility))}

    @app.post("/v1/kb/{kb_id}/delete")
    async def delete_kb_ep(kb_id: str, caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        return {"deleted": bool(state.kb_service.delete_kb(caller, kb_id))}

    # ---- 知识库批量入库 + 类型映射表 + 配置 + 自定义条目 (v0.7.1) ----
    from backend.kb_batch_ingest import ingest_batch, IngestConfig
    from backend.kb_ingest_wiring import REGISTRY, load_config, save_config, make_extract_fn
    from backend.kb_custom_fields import CustomFieldService, FieldType

    _FIELDS = CustomFieldService()

    @app.post("/v1/kb/{kb_id}/batch-ingest")
    async def kb_batch_ingest_ep(kb_id: str, files: "List[UploadFile]" = File(...),
                                 caller: "User" = Depends(get_caller)) -> Dict[str, Any]:
        cfg = load_config()
        uploads = [(f.filename or "upload", await f.read()) for f in files]
        def _ingest(kbid, doc_id, text):
            state.kb_service.ingest_text(caller, kbid, doc_id, text)
        rep = ingest_batch(uploads, kb_id=kb_id, registry=REGISTRY,
                           extract_fn=make_extract_fn(cfg), ingest_fn=_ingest, cfg=cfg)
        return rep.public()

    @app.get("/v1/kb/parsers")
    async def list_parsers_ep(_: "User" = Depends(get_caller)) -> list:
        return [m.__dict__ for m in REGISTRY.list()]

    class ParserAdd(BaseModel):
        ext: str; parser: str = ""; external_cmd: str = ""; note: str = ""

    @app.post("/v1/kb/parsers")
    async def add_parser_ep(req: "ParserAdd", _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        try:
            return REGISTRY.add(req.ext, parser=req.parser, external_cmd=req.external_cmd, note=req.note).__dict__
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    class ParserImport(BaseModel):
        rows: list

    @app.post("/v1/kb/parsers/import")
    async def import_parsers_ep(req: "ParserImport", _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        return {"imported": REGISTRY.import_rows(req.rows)}

    @app.get("/v1/kb/ingest-config")
    async def get_ingest_cfg_ep(_: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        return load_config().__dict__

    class IngestCfgReq(BaseModel):
        max_files: int = 50; max_file_mb: int = 20; max_total_mb: int = 100
        ocr_enabled: bool = False; ocr_lang: str = "ch"

    @app.post("/v1/kb/ingest-config")
    async def set_ingest_cfg_ep(req: "IngestCfgReq", _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        cfg = IngestConfig(**req.model_dump()); save_config(cfg); return cfg.__dict__

    class FieldSuggestReq(BaseModel):
        name: str; sample: str = ""

    @app.post("/v1/kb/fields/suggest")
    async def suggest_field_ep(req: "FieldSuggestReq", _: "User" = Depends(get_caller)) -> Dict[str, Any]:
        s = _FIELDS.suggest(req.name, req.sample)
        return {"name": s.name, "suggested_type": s.suggested_type.value,
                "suggested_value": s.suggested_value, "rationale": s.rationale}

    class FieldConfirmReq(BaseModel):
        name: str; type: str

    @app.post("/v1/kb/fields/confirm")
    async def confirm_field_ep(req: "FieldConfirmReq", _: "User" = Depends(get_caller)) -> Dict[str, Any]:
        try:
            d = _FIELDS.confirm(req.name, FieldType(req.type))
            return {"name": d.name, "type": d.type.value, "confirmed": True}
        except ValueError:
            raise HTTPException(status_code=400, detail="类型须为 bool/number/text/date")

    # ---- 白标:管理后台可维护 (v0.7.2) ----
    from backend.branding_store import Branding, get_branding, set_branding

    @app.get("/v1/branding")
    async def get_branding_ep() -> Dict[str, Any]:
        return get_branding().public()

    class BrandingIn(BaseModel):
        platform_name: Optional[str] = None
        logo_url: Optional[str] = None
        favicon_url: Optional[str] = None
        brand_color: Optional[str] = None
        brand_color_dark: Optional[str] = None
        lock_accent: Optional[bool] = None
        login_tagline: Optional[str] = None

    @app.put("/v1/branding")
    async def put_branding_ep(req: "BrandingIn",
                              _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        b = get_branding()
        for k, v in req.model_dump(exclude_none=True).items():
            setattr(b, k, v)
        if b.logo_url.startswith("data:") and len(b.logo_url) > 700_000:
            raise HTTPException(status_code=400, detail="logo 过大,请压到 ~500KB 内或用图片URL")
        set_branding(b)
        return b.public()

    
    # ---------------- 对话（可选认证：匿名可用，带 token 则注入身份） ----------------
    @app.websocket("/v1/chat/stream")
    async def chat_stream(ws: WebSocket) -> None:
        """流式对话。收 {message, agent_id, session_id}，逐帧推 trace/delta/done。"""
        await ws.accept()
        try:
            req = await ws.receive_json()
            token = req.get("token")
            caller = state.auth.resolve(token)
            caller_d = {"id": caller.id, "role": caller.role.value, "dept_id": caller.dept_id} if caller else None
            async for ev in run_chat(
                state,
                message=req.get("message", ""),
                agent_id=req.get("agent_id", "general"),
                session_id=req.get("session_id", "anon"),
                caller=caller_d,
                ref_sessions=req.get("ref_sessions"),
                model=req.get("model"),
            ):
                await ws.send_json(ev)
        except WebSocketDisconnect:
            return
        except Exception as e:  # 任何异常都干净收尾，不泄露内部细节
            await ws.send_json({"event": "done", "blocked": True, "reason": "server_error"})
            _ = e
        finally:
            await ws.close()

    @app.post("/v1/chat")
    async def chat(req: "ChatRequest",
                   caller: "Optional[User]" = Depends(get_caller_optional)) -> Dict[str, Any]:
        """非流式回退：聚合流式事件为一次性结果。"""
        caller_d = {"id": caller.id, "role": caller.role.value, "dept_id": caller.dept_id} if caller else None
        events = run_chat(state, message=req.message, agent_id=req.agent_id,
                          session_id=req.session_id, caller=caller_d)
        return await collect_answer(events)
    # ── 上安 v2.0 集成: Kun/Kimi/Artifacts 中间件 ──
    from .shangan_integration.pipeline import api as shangan_api
    app.include_router(shangan_api)

        # ==== v0.9.0: 整站搬迁(管理员) ====
    import tempfile, time, os as _os2
    from .platform_storage import export_site, import_site, read_manifest
    from .platform_storage.config import data_dir

    @app.post("/v1/admin/site/export")
    async def site_export_ep(_: "User" = Depends(require_role(Role.ADMIN))):
        out = _os2.path.join(tempfile.gettempdir(), f"site-{int(time.time())}.tar.gz")
        export_site(data_dir(), out, platform_version=_os2.getenv("PLATFORM_VERSION", ""))
        return FileResponse(out, media_type="application/gzip", filename="site-backup.tar.gz")

    @app.post("/v1/admin/site/import")
    async def site_import_ep(file: "UploadFile" = File(...),
                             overwrite: bool = Form(False),
                             _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        tmp = _os2.path.join(tempfile.gettempdir(), "site-import.tar.gz")
        with open(tmp, "wb") as f:
            f.write(await file.read())
        try:
            info = read_manifest(tmp)
            res = import_site(tmp, data_dir(), overwrite=overwrite)
            return {"ok": True, "manifest": info, "result": res,
                    "note": "导入后请重启服务以加载新数据"}
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    @app.post("/v1/admin/site/inspect")
    async def site_inspect_ep(file: "UploadFile" = File(...),
                              _: "User" = Depends(require_role(Role.ADMIN))) -> Dict[str, Any]:
        tmp = _os2.path.join(tempfile.gettempdir(), "site-inspect.tar.gz")
        with open(tmp, "wb") as f:
            f.write(await file.read())
        try:
            return read_manifest(tmp)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))


    # ---- 前端静态文件（SPA 回退）----
    import os
    from fastapi.staticfiles import StaticFiles
    from fastapi.responses import FileResponse
    FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "claude-frontend", "frontend-new", "dist")
    if os.path.isdir(FRONTEND_DIR):
        app.mount("/assets", StaticFiles(directory=os.path.join(FRONTEND_DIR, "assets")), name="assets")
        @app.get("/{full_path:path}")
        async def spa_fallback(full_path: str):
            fp = os.path.join(FRONTEND_DIR, full_path)
            if full_path and os.path.isfile(fp):
                return FileResponse(fp)
            return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))

else:
    app = None  # 离线占位
